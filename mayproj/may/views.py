from email.mime.text import MIMEText

from django.core.mail import EmailMultiAlternatives
from django.db import DataError
from django.shortcuts import render
import io
import requests
import time
import datetime
from django.contrib.sessions.models import Session
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib.auth.models import User, AnonymousUser
from django.contrib.auth import login, authenticate
from django.conf import settings
from django import forms
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, HttpResponseNotAllowed
from django.utils.datastructures import MultiValueDictKeyError
from django.views.decorators.cache import cache_control
from time import gmtime, strftime

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.fields import empty
from rest_framework.permissions import AllowAny

# Create your views here.
from may.models import Member, ComMember, Talker, Group, Note, Attendee, TalkingRoom


class UploadFileForm(forms.Form):
    file = forms.FileField()

def index(request):
    # return HttpResponse('<h1>Hello Cay Sees!</h1>')
    if not request.session.exists(request.session.session_key):
        request.session.create()
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    idx = session.get_decoded().get('_auth_user_id')
    users = Member.objects.filter(admin_id=idx)
    if users.count() > 0:
        email = users[0].email
        pwd = users[0].password
        user = authenticate(email=email, password=pwd)
        login(request, user)
        request.session['user_id'] = users[0].pk
        if users[0].role != 'admin':
            return redirect('/createTalkingLink')
        else:
            return redirect('/logout')
    else:
        try:
            talker = Talker.objects.filter(id=request.session['talker_id'])
            users = ComMember.objects.filter(email=talker.email)
            if users.count() > 0:
                return redirect('/createTalkingLink')
            else:
                return redirect('/logout')
        except KeyError:
            return redirect('/logout')
        except AttributeError:
            return redirect('/logout')

@cache_control(no_cache=True, must_revalidate=True)
def logout(request):
    from django.contrib.auth import logout
    from django.contrib import auth
    from django.core.cache import cache

    try:
        if request.session['user'] == 'admin':
            request.session['readterms'] = ''
            request.session['user_id'] = ''
            request.session['talker_id'] = ''
            request.session['user'] = ''
            request.session['company'] = ''
            logout(request)
            cache.clear()
            request.session.flush()
            request.user = AnonymousUser
            context = {'email': ''}
            return redirect('/manager')
        else:
            request.session['readterms'] = ''
            request.session['user_id'] = ''
            request.session['talker_id'] = ''
            request.session['user'] = ''
            request.session['company'] = ''
            logout(request)
            cache.clear()
            request.session.flush()
            request.user = AnonymousUser
            context = {'email': ''}
            return render(request, 'may/origin.html')
    except KeyError:
        request.session['readterms'] = ''
        request.session['user_id'] = ''
        request.session['talker_id'] = ''
        request.session['user'] = ''
        request.session['company'] = ''
        logout(request)
        cache.clear()
        request.session.flush()
        request.user = AnonymousUser
        context = {'email': ''}
        return render(request, 'may/origin.html')

def tologin(request):
    request.session['user'] = 'individual'
    return render(request, 'may/login.html')

def tocomlogin(request):
    request.session['user'] = 'company'
    return render(request, 'may/comlogin.html')

def tosignup(request):
    return render(request, 'may/signup.html')

def torequestpwd(request):
    return render(request, 'may/forgotpwd.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def toterms(request):
    if request.method == 'POST':
        eml = request.POST.get('email', '')
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if request.session['user'] == 'individual':
            users = Member.objects.filter(email=eml)
            if users.count() > 0:
                if users[0].terms != 'read':
                    return render(request, 'may/terms.html')
                else:
                    return render(request, 'may/result.html',
                                  {'response': 'You have already registered. Please login...'})
            else:
                return render(request, 'may/terms.html')
        elif request.session['user'] == 'company':
            users = ComMember.objects.filter(email=eml)
            if users.count() > 0:
                if users[0].terms != 'read':
                    return render(request, 'may/terms.html')
                else:
                    return render(request, 'may/result.html',
                                  {'response': 'You have already read the terms & conditions'})
            else:
                return render(request, 'may/terms.html')

def termsdone(request):
    request.session['readterms'] = 'true'
 #   return HttpResponse(request.session['readterms'])
    return render(request, 'may/termsdone.html')

def terms2(request):
    return render(request, 'may/terms2.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def welcome(request):
    if request.method == 'POST':
        eml = request.POST.get('email', None)
        name = request.POST.get('name', None)
        password = request.POST.get('password', None)
        usrs = Talker.objects.filter(email=eml)
        if usrs.count() > 0:
            return render(request, 'may/result.html',
                          {'response': 'Someone already is using the same email.'})
        users = Member.objects.filter(email=eml)
        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')
        if request.session['readterms'] != 'true':
            return render(request, 'may/result.html',
                          {'response': 'Please read terms & conditions'})
        count = users.count()
        if count == 0:
            user = Member()
            user.email = eml
            user.name = name
            user.password = password
            user.terms = 'read'
            user.role = request.session['user']
            user.company = ''

            user.save()

            talkers = Talker.objects.filter(email=eml)
            if talkers.count() == 0:
                talker = Talker()
                talker.name = name
                talker.email = eml
                talker.password = password
                talker.company = ''
                talker.callreg = '0'
                talker.save()

            user1 = User()
            user1.username = eml
            user1.email = eml
            user1.password = password
            user1.set_password(password)
            user1.save()

            user.admin_id = user1.pk
            user.save()

            user1 = authenticate(username=eml, password=password)
            login(request, user1)
            request.session['user_id'] = user.pk
            talkers = Talker.objects.filter(email=eml)
            request.session['talker_id'] = talkers[0].pk
            return render(request, 'may/welcome.html')
        else:
            return render(request, 'may/result.html',
                          {'response': 'You have already been registered. Please log in...'})

def home(request):
    try:
        user_id = request.session['talker_id']
        try:
            rooms = TalkingRoom.objects.filter(talker1=user_id, talker2=request.session['the'])
            rooms[0].delete()
        except KeyError:
            print('No user talked to')
        except IndexError:
            print('No user talked to')
        me = Talker.objects.get(id=user_id)
        if request.session['user'] == 'admin':
            return redirect('/adminhome')
        else:
            return render(request, 'may/home.html', {'me': me})
    except KeyError:
        return redirect('/logout')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login_user(request):
    if request.method == 'POST':
        email = request.POST.get('email', None)
        password = request.POST.get('password', None)

        users = Member.objects.filter(email=email, password=password)
        count = users.count()
        if count > 0:
            user = authenticate(username=email, password=password)
            login(request, user)
            # return HttpResponse(user.email)
            if user is not None:
                request.session['user_id'] = users[0].pk
                if users[0].role != 'admin':
                    try:
                        if users[0].role == request.session['user']:
                            talkers = Talker.objects.filter(email=email)
                            request.session['talker_id'] = talkers[0].pk
                            return redirect('/createTalkingLink')
                        else:
                            logOut(request)
                            return render(request, 'may/result.html',
                                          {'response': 'You have not been registered. Please register...'})
                    except KeyError:
                        logOut(request)
                        return render(request, 'may/result.html',
                                      {'response': 'You have not been registered. Please register...'})
                else:
                    talkers = Talker.objects.filter(email=email)
                    request.session['talker_id'] = talkers[0].pk
                    request.session['user'] = 'admin'
                    request.session['company'] = talkers[0].company
                    return redirect('/createTalkingLink')
            else:
                logOut(request)
                return render(request, 'may/result.html',
                              {'response': 'You have not been registered. Please register...'})
        else:
            return render(request, 'may/result.html',
                          {'response': 'You have not been registered. Please register...'})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def login_comuser(request):
    if request.method == 'POST':
        email = request.POST.get('email', None)
        password = request.POST.get('password', None)

        users = ComMember.objects.filter(email=email, password=password)
        count = users.count()
        if count > 0:
            user = users[0]
            try:
                if request.session['readterms'] == 'true':
                    user.terms = 'read'
                    user.save()
                elif user.terms != 'read':
                    return render(request, 'may/result.html',
                                  {'response': 'Please read terms & conditions...'})
            except KeyError:
                if user.terms != 'read':
                    return render(request, 'may/result.html',
                                  {'response': 'Please read terms & conditions...'})
            request.session['comuser_id'] = user.pk
            talkers = Talker.objects.filter(email=email)
            if talkers.count() == 0:
                talker = Talker()
                talker.name = user.name
                talker.email = user.email
                talker.password = password
                talker.company = user.company
                talker.callreg = '0'
                talker.save()
            talkers = Talker.objects.filter(email=email)
            request.session['talker_id'] = talkers[0].pk
            return redirect('/createTalkingLink')
        else:
            return render(request, 'may/result.html',
                          {'response': 'You have not been registered. Please register...'})

def logOut(request):
    from django.contrib.auth import logout
    from django.contrib import auth
    from django.core.cache import cache
    logout(request)
    cache.clear()
    request.session.flush()

def myprofile(request):
    user = Member.objects.get(id=request.session['user_id'])
    user_id = request.session['talker_id']
    users = ComMember.objects.filter(adminID=request.session['user_id']).order_by('-id')
    context = {
        'members': users,
        'admin': user,
        'user_id':user_id
    }
    return render(request, 'may/adminhome.html', context)

def profile(request):
    me = Talker.objects.get(id=request.session['talker_id'])
    context = {
        'me': me,
        'profile': 'true'
    }
    return render(request, 'may/home.html', context)

def toadminlogin(request):
    return render(request,'may/adminlogin.html')

def toadminsignup(request):
    return render(request,'may/adminsignup.html')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def adminsignup(request):
    if request.method == 'POST':
        eml = request.POST.get('email', None)
        name = request.POST.get('name', None)
        password = request.POST.get('password', None)
        company = request.POST.get('company', None)
        usrs = Talker.objects.filter(email=eml)
        if usrs.count() > 0:
            return render(request, 'may/result.html',
                          {'response': 'Someone already is using the same email.'})
        users = Member.objects.filter(email=eml)
        count = users.count()
        if count == 0:
            user = Member()
            user.email = eml
            user.name = name
            user.password = password
            user.role = 'admin'
            user.company = company

            user.save()

            talkers = Talker.objects.filter(email=eml)
            if talkers.count() == 0:
                talker = Talker()
                talker.name = name
                talker.email = eml
                talker.password = password
                talker.company = company
                talker.callreg = '0'
                talker.save()

            talkers = Talker.objects.filter(email=eml)
            request.session['talker_id'] = talkers[0].pk

            user1 = User()
            user1.username = eml
            user1.email = eml
            user1.password = password
            user1.set_password(password)
            user1.save()

            user.admin_id = user1.pk
            user.save()
            user1 = authenticate(username=eml, password=password)
            login(request, user1)
            request.session['user_id'] = users[0].pk
            request.session['company'] = company
            request.session['user'] = 'admin'
            return redirect('/createTalkingLink')
        else:
            return render(request, 'may/result.html',
                          {'response': 'You have already been registered. Please log in...'})

def adminhome(request):
    users = []
    user_id = request.session['talker_id']
    request.session['user'] = 'admin'
    users = ComMember.objects.filter(adminID=request.session['user_id']).order_by('-id')
    context = {
        'members':users,
        'user_id': user_id
    }
    return render(request, 'may/adminhome.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def addmember(request):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        usrs = Talker.objects.filter(email=email)
        if usrs.count() > 0:
            return render(request, 'may/result.html',
                          {'response': 'Someone already is using the same email.'})
        comusers = ComMember.objects.filter(email=email)
        if comusers.count() == 0:
            comuser = ComMember()
            comuser.adminID = request.session['user_id']
            comuser.name = name
            comuser.email = email
            comuser.password = password
            comuser.company = request.session['company']
            comuser.save()

            me = Talker.objects.get(id=request.session['talker_id'])
            fromaddress = me.email
            mailList = []
            mailList.append(email)

            message = 'You are allowed to login with scwribe mobile website to use it.<br>Your username: ' + email + ', and your password: ' + password + '<br><br>' + me.company + '<br><br>Enjoy scwribe talkig<br>https://scwribe.com'

            html = """\
                                    <html>
                                      <head></head>
                                      <body>
                                      <a href="https://scwribe.com"><img src="https://scwribe.com/static/may/images/logo.png" style="width:90px;height:90px;border-radius: 8%; margin-left:25px;"/></a>
                                        <h2 style="margin-left:10px; color:#02839a;">Company's approvement for it's member to use scwribe mobile website</h2>
                                        <div style="font-size:16px; word-break: break-all; word-wrap: break-word;">
                                            {mes}
                                        </div>
                                      </body>
                                    </html>
                                    """
            html = html.format(mes=message)

            msg = EmailMultiAlternatives('We allowed you to use scwribe website for good talking', '', fromaddress, mailList)
            msg.attach_alternative(html, "text/html")
            msg.send(fail_silently=False)

            return redirect('/adminhome')
        else:
            return render(request, 'may/result.html',
                          {'response': 'This email exists already. Try again...'})
    else:
        return redirect('/adminhome')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def deletemember(request):
    if request.method == 'POST':
        memberids = request.POST.getlist('items[]')
        for i in memberids:
            ComMember.objects.get(id=i).delete()
        return redirect('/adminhome')
    else:
        return redirect('/adminhome')

def editbox(request, member_id):
    user_id = request.session['user_id']
    comuser = ComMember.objects.get(id=member_id)
    users = ComMember.objects.filter(adminID=request.session['user_id']).order_by('-id')
    context = {
        'members': users,
        'editmember': comuser,
        'user_id': user_id
    }
    return render(request, 'may/adminhome.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updatemember(request, member_id):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')

        comuser = ComMember.objects.get(id=member_id)
        comuser.name = name
        comuser.email = email
        comuser.password = password
        comuser.save()
        return redirect('/adminhome')
    else:
        return redirect('/adminhome')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updateadmin(request, admin_id):
    if request.method == 'POST':
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        password = request.POST.get('password', '')
        company = request.POST.get('company', '')

        if not request.session.exists(request.session.session_key):
            request.session.create()
        session_key = request.session.session_key
        session = Session.objects.get(session_key=session_key)
        idx = session.get_decoded().get('_auth_user_id')

        user = Member.objects.get(id=admin_id)
        user.email = email
        user.name = name
        pwd = user.password
        user.password = password
        user.role = 'admin'
        user.company = company

        user1 = authenticate(username=email, password=pwd)
        if user1 is not None:
            user1.password = password
            user1.set_password(password)
            user1.save()

            user.save()

            talkers = Talker.objects.filter(email=email)
            talker = talkers[0]
            talker.password = password
            talker.save()

        request.session['user_id'] = user.pk

        return redirect('/createTalkingLink')
    else:
        return redirect('/adminhome')

@csrf_protect
@csrf_exempt
def searchcomuser(request):
    if request.method == 'POST':

        search_id = request.POST.get('q', None)
        user_id = request.session['talker_id']
        users = ComMember.objects.filter(name__icontains=search_id, adminID=request.session['user_id']).order_by('-id')

        if users.count()>0:
            context = {'members': users, 'user_id':user_id}
            return render(request, 'may/adminhome.html', context)

        else:
            users = ComMember.objects.filter(email__icontains=search_id, adminID=request.session['user_id']).order_by('-id')
            context = {'members': users, 'user_id': user_id}
            return render(request, 'may/adminhome.html', context)

    else:
        return redirect('/adminhome')

def addbox(request):
    user_id = request.session['talker_id']
    users = ComMember.objects.filter(adminID=request.session['user_id']).order_by('-id')
    context = {
        'members': users,
        'addbox': 'true',
        'user_id': user_id
    }
    return render(request, 'may/adminhome.html', context)

def export_xlsx_member(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=member_template_xlsx.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Company Members"

    row_num = 0

    columns = [
        (u"Name", 20),
        (u"Email", 30),
        (u"Password", 30),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]

        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    wb.save(response)
    return response

def import_view_member(request):

    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)
    else:
        form = UploadFileForm()
    return render(
        request,
        'may/upload_form.html',
        {
            'form': form,
            'title': 'Load Data',
            'header': 'Upload company member info:'
        })

import xlrd
from xlrd import XLRDError
from _mysql_exceptions import DataError

def import_member_data(request):

    if request.method == "POST":
        form = UploadFileForm(request.POST,
                              request.FILES)

        if form.is_valid():

            input_excel = request.FILES['file']
            try:
                book = xlrd.open_workbook(file_contents=input_excel.read())
                sheet = book.sheet_by_index(0)

                for r in range(1, sheet.nrows):
                    name = sheet.cell(r, 0).value
                    email = sheet.cell(r, 1).value
                    password = sheet.cell(r, 2).value

                    users = ComMember.objects.filter(email=email)
                    if users.count()==0:
                        user = ComMember()
                        user.adminID = request.session['user_id']
                        user.name = name
                        user.email = email
                        user.password = password

                        user.save()

                return redirect('/adminhome')

            except XLRDError:
                return render(request, 'may/upload_form.html', {'note':'invalid_file'})
            except IOError:
                return render(request, 'may/upload_form.html', {'note': 'invalid_file'})
            except IndexError:
                return render(request, 'may/upload_form.html', {'note': 'invalid_file'})
            except DataError:
                return HttpResponse('Invalid file!')
        else:
            return render(request, 'may/upload_form.html', {'note':'invalid_file'})

    else:
        return redirect('/adminhome')

def meet(request, user_id):
    roomlink = ''
    me = Talker.objects.get(id=request.session['talker_id'])
    user = Talker.objects.get(id=user_id)
    rooms = TalkingRoom.objects.filter(talker1=user_id, talker2=request.session['talker_id'])
    if rooms.count() > 0:
        roomlink = rooms[0].roomlink
    else:
        room = TalkingRoom()
        room.talker1 = request.session['talker_id']
        room.talker2 = user_id
        room.roomlink = me.link
        room.save()
        roomlink = me.link
    context = {
        'sender': me,
        'user': user,
        'roomlink': roomlink
    }
    request.session['the'] = user_id
    return render(request, 'may/chattest.html', context)

def users(request):
    talkers = []
    users = Talker.objects.all()
    me = Talker.objects.get(id=request.session['talker_id'])
    for user in users:
        if user.email != me.email:
            talkers.insert(0, user)
    groups = Group.objects.filter(talker_id=request.session['talker_id']).order_by('-id')
    return render(request, 'may/users.html', {'users':talkers, 'groups':groups})

def allgroups(request):
    groups = Group.objects.all().order_by('-id')
    return render(request, 'may/groups.html', {'groups':groups})

def addgroupbox(request):
    groups = Group.objects.all()
    return render(request, 'may/groups.html', {'groups':groups, 'addgroup': 'true'})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def creategroup(request):
    if request.method == 'POST':
        groupname = request.POST.get('name', '')
        me = Talker.objects.get(id=request.session['talker_id'])
        group = Group()
        group.talker_id = me.pk
        group.talker_name = me.name
        group.name = groupname
        group.date = strftime("%Y-%m-%d %H:%M", gmtime())
        group.save()
        groups = Group.objects.all()
        return render(request, 'may/groups.html', {'groups':groups})
    else:
        return redirect('/groups')

def mygroups(request):
    talkers = []
    users = Talker.objects.all()
    me = Talker.objects.get(id=request.session['talker_id'])
    for user in users:
        if user.email != me.email:
            talkers.insert(0, user)
    groups = Group.objects.filter(talker_id=request.session['talker_id']).order_by('-id')
    return render(request, 'may/mygroups.html', {'mygroups':groups, 'users':talkers})

def editgroupbox(request, group_id):
    group = Group.objects.get(id=group_id)
    groups = Group.objects.filter(talker_id=request.session['talker_id'])
    context = {
        'mygroups': groups,
        'group': group,
        'editgroup':'true'
    }
    return render(request, 'may/mygroups.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def updategroup(request, group_id):
    if request.method == 'POST':
        name = request.POST.get('name', '')

        group = Group.objects.get(id=group_id)
        group.name = name
        group.date = strftime("%Y-%m-%d %H:%M", gmtime())

        group.save()
        return redirect('/mygroups')
    else:
        return redirect('/mygroups')

def deletegroup(request, group_id):
    Group.objects.get(id=group_id).delete()
    return redirect('/mygroups')

def meetgroup(request, group_id):
    me = Talker.objects.get(id=request.session['talker_id'])
    group = Group.objects.get(id=group_id)
    createrid = group.talker_id
    creater = Talker.objects.get(id=createrid)
    roomlink = creater.link
    context = {
        'sender': me,
        'group': group,
        'roomlink': roomlink
    }
    return render(request, 'may/meet.html', context)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def transcript(request, param):
    if request.method == 'POST':
        talker = request.POST.get('talker', '')
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        company = request.POST.get('company', '')
        note1 = request.POST.get('note', '')
        attendeeids = request.POST.getlist('attendeeids[]')
        attendeenotes = request.POST.getlist('attendeenotes[]')
        lang = request.POST.get('lang', '')
        startTime = request.POST.get('startTime', '')
        endTime = request.POST.get('endTime', '')
        urgency = request.POST.get('urgency', '')
        topic = request.POST.get('topic', '')
        summary = request.POST.get('summary', '')
        callofaction = request.POST.get('callofaction', '')

        note = Note()
        note.talker = talker
        note.name = name
        note.email = email
        note.company = company
        note.note = note1
        note.lang = lang
        note.start_time = startTime
        note.end_time = endTime
        note.urgency = urgency
        note.topic = topic
        note.summary = summary
        note.callofaction = callofaction
        note.save()

        for i in attendeeids:
            attendee = Attendee()
            attendee.note_id = note.pk
            attendee.talker = i
            user = Talker.objects.get(id=i)
            attendee.name = user.name
            attendee.email = user.email
            attendee.company = user.company
            try:
                attendee.note = attendeenotes[int(i)]
            except IndexError:
                attendee.note = ''
            except Exception:
                attendee.note = ''
            attendee.lang = lang
            attendee.save()

        if param == '0':
            return redirect('/meet/' + attendeeids[0])
        elif param == '1':
            if request.session['user'] == 'admin':
                return redirect('/adminhome')
            else:
                return redirect('/users')
    else:
        if request.session['user'] == 'admin':
            return redirect('/adminhome')
        else:
            return redirect('/users')

def transcripts(request, user_id):
    noteList = []
    attendeeList = []
    list = []
    notes = Note.objects.filter(talker=request.session['talker_id'])
    for note in notes:
        attendees = Attendee.objects.filter(talker = user_id, note_id = note.pk)
        noteList.insert(0, note)
        attendeeList.insert(0, attendees)

    return render(request, 'may/transcripts.html', {'notes':noteList, 'attendees':attendeeList, 'user_id':user_id})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def deletenote(request, user_id):
    if request.method == 'POST':
        tids = request.POST.getlist('items[]')
        for i in tids:
            Note.objects.get(id=i).delete()
            attendees = Attendee.objects.filter(note_id=i)
            for attendee in attendees:
                attendee.delete()
        return redirect('/transcripts/' + user_id)
    else:
        return redirect('/transcripts/' + user_id)

def notedetail(request, note_id, user_id):
    note = Note.objects.get(id=note_id)
    attendees = Attendee.objects.filter(talker=user_id, note_id=note.pk)
    return render(request, 'may/transcriptdetail.html', {'note':note, 'attendees':attendees, 'user_id':user_id})


@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def grouptranscript(request, param):
    if request.method == 'POST':
        talker = request.POST.get('talker', '')
        name = request.POST.get('name', '')
        email = request.POST.get('email', '')
        company = request.POST.get('company', '')
        note1 = request.POST.get('note', '')
        attendeeids = request.POST.getlist('attendeeids[]')
        attendeenotes = request.POST.getlist('attendeenotes[]')
        lang = request.POST.get('lang', '')
        startTime = request.POST.get('startTime', '')
        endTime = request.POST.get('endTime', '')
        urgency = request.POST.get('urgency', '')
        topic = request.POST.get('topic', '')
        summary = request.POST.get('summary', '')
        callofaction = request.POST.get('callofaction', '')
        groupid = request.POST.get('groupid', '')

        note = Note()
        note.talker = talker
        note.name = name
        note.email = email
        note.company = company
        note.note = note1
        note.lang = lang
        note.group = groupid
        note.start_time = startTime
        note.end_time = endTime
        note.urgency = urgency
        note.topic = topic
        note.summary = summary
        note.callofaction = callofaction
        note.save()

        for i in attendeeids:
            attendee = Attendee()
            attendee.note_id = note.pk
            attendee.talker = i
            user = Talker.objects.get(id=i)
            attendee.name = user.name
            attendee.email = user.email
            attendee.company = user.company
            try:
                attendee.note = attendeenotes[int(i)]
            except IndexError:
                attendee.note = ''
            except Exception:
                attendee.note = ''
            attendee.lang = lang
            attendee.save()

        if param == '0':
            return redirect('/meetgroup/' + groupid)
        elif param == '1':
            return redirect('/home')
    else:
        return redirect('/home')

def allnotes(request):
    noteList = []
    attendeeList = []
    list = []
    notes = Note.objects.all()
    for note in notes:
        if note.group != '':
            attendees = Attendee.objects.filter(note_id = note.pk)
            noteList.insert(0, note)
            attendeeList.insert(0, attendees)

    return render(request, 'may/allnotes.html', {'notes':noteList, 'attendees':attendeeList})

def transcriptdetail(request, note_id):
    note = Note.objects.get(id=note_id)
    attendees = Attendee.objects.filter(note_id=note.pk)
    context = {'note':note, 'attendees':attendees}
    return render(request, 'may/notedetail.html', context)
    # if note.group != '':
    #     return render(request, 'may/notedetail.html', context)
    # else:
    #     return render(request, 'may/notedetail2.html', context)

@csrf_protect
@csrf_exempt
def searchnotes(request):
    if request.method == 'POST':
        search_id = request.POST.get('q', None)
        notes = Note.objects.filter(name__icontains=search_id).order_by('-id')

        if notes.count()>0:
            noteList = []
            attendeeList = []
            for note in notes:
                if note.group != '':
                    attendees = Attendee.objects.filter(note_id=note.pk)
                    noteList.insert(0, note)
                    attendeeList.insert(0, attendees)

            return render(request, 'may/allnotes.html', {'notes': noteList, 'attendees': attendeeList})

        else:
            notes = Note.objects.filter(email__icontains=search_id).order_by('-id')
            if notes.count() > 0:
                noteList = []
                attendeeList = []
                for note in notes:
                    if note.group != '':
                        attendees = Attendee.objects.filter(note_id=note.pk)
                        noteList.insert(0, note)
                        attendeeList.insert(0, attendees)

                return render(request, 'may/allnotes.html', {'notes': noteList, 'attendees': attendeeList})
            else:
                notes = Note.objects.filter(note__icontains=search_id).order_by('-id')
                if notes.count() > 0:
                    noteList = []
                    attendeeList = []
                    for note in notes:
                        if note.group != '':
                            attendees = Attendee.objects.filter(note_id=note.pk)
                            noteList.insert(0, note)
                            attendeeList.insert(0, attendees)

                    return render(request, 'may/allnotes.html', {'notes': noteList, 'attendees': attendeeList})
                else:
                    notes = Note.objects.filter(urgency__icontains=search_id).order_by('-id')
                    if notes.count() > 0:
                        noteList = []
                        attendeeList = []
                        for note in notes:
                            if note.group != '':
                                attendees = Attendee.objects.filter(note_id=note.pk)
                                noteList.insert(0, note)
                                attendeeList.insert(0, attendees)

                        return render(request, 'may/allnotes.html', {'notes': noteList, 'attendees': attendeeList})
                    else:
                        notes = Note.objects.filter(summary__icontains=search_id).order_by('-id')
                        if notes.count() > 0:
                            noteList = []
                            attendeeList = []
                            for note in notes:
                                if note.group != '':
                                    attendees = Attendee.objects.filter(note_id=note.pk)
                                    noteList.insert(0, note)
                                    attendeeList.insert(0, attendees)

                            return render(request, 'may/allnotes.html', {'notes': noteList, 'attendees': attendeeList})
                        else:
                            notes = Note.objects.filter(callofaction__icontains=search_id).order_by('-id')
                            if notes.count() > 0:
                                noteList = []
                                attendeeList = []
                                for note in notes:
                                    if note.group != '':
                                        attendees = Attendee.objects.filter(note_id=note.pk)
                                        noteList.insert(0, note)
                                        attendeeList.insert(0, attendees)

                                return render(request, 'may/allnotes.html',
                                              {'notes': noteList, 'attendees': attendeeList})
                            else:
                                notes = Note.objects.filter(email__icontains=search_id).order_by('-id')
                                if notes.count() > 0:
                                    noteList = []
                                    attendeeList = []
                                    for note in notes:
                                        if note.group != '':
                                            attendees = Attendee.objects.filter(note_id=note.pk)
                                            noteList.insert(0, note)
                                            attendeeList.insert(0, attendees)

                                    return render(request, 'may/allnotes.html',
                                                  {'notes': noteList, 'attendees': attendeeList})
                                else:
                                    allnotes = Note.objects.all()
                                    noteList = []
                                    attendeeList = []
                                    for note in allnotes:
                                        st = time.strftime('%Y-%m-%d %H:%M', time.gmtime(int(note.start_time) / 1000.0))
                                        et = time.strftime('%Y-%m-%d %H:%M', time.gmtime(int(note.end_time) / 1000.0))
                                        if search_id in st:
                                            attendees = Attendee.objects.filter(note_id=note.pk)
                                            noteList.insert(0, note)
                                            attendeeList.insert(0, attendees)
                                        elif search_id in et:
                                            attendees = Attendee.objects.filter(note_id=note.pk)
                                            noteList.insert(0, note)
                                            attendeeList.insert(0, attendees)
                                    return render(request, 'may/allnotes.html',
                                                  {'notes': noteList, 'attendees': attendeeList})

    else:
        return redirect('/allnotes')

@csrf_protect
@csrf_exempt
def schusers(request):
    if request.method == 'POST':
        search_id = request.POST.get('q', None)
        users = Talker.objects.filter(name__icontains=search_id).order_by('-id')
        if users.count()>0:
            talkers = []
            me = Talker.objects.get(id=request.session['talker_id'])
            for user in users:
                if user.email != me.email:
                    talkers.insert(0, user)
            return render(request, 'may/users.html', {'users': talkers})
        else:
            users = Talker.objects.filter(email__icontains=search_id).order_by('-id')
            if users.count() > 0:
                talkers = []
                me = Talker.objects.get(id=request.session['talker_id'])
                for user in users:
                    if user.email != me.email:
                        talkers.insert(0, user)
                return render(request, 'may/users.html', {'users': talkers})
            else:
                users = Talker.objects.filter(company__icontains=search_id).order_by('-id')
                if users.count() > 0:
                    talkers = []
                    me = Talker.objects.get(id=request.session['talker_id'])
                    for user in users:
                        if user.email != me.email:
                            talkers.insert(0, user)
                    return render(request, 'may/users.html', {'users': talkers})
                else:
                    return redirect('/users')

    else:
        return redirect('/users')

@csrf_protect
@csrf_exempt
def schgroups(request):
    if request.method == 'POST':
        search_id = request.POST.get('q', None)
        groups = Group.objects.filter(talker_name__icontains=search_id, talker_id=request.session['talker_id']).order_by('-id')
        if groups.count()>0:
            return render(request, 'may/mygroups.html', {'mygroups': groups})
        else:
            groups = Group.objects.filter(name__icontains=search_id, talker_id=request.session['talker_id']).order_by('-id')
            if groups.count() > 0:
                return render(request, 'may/mygroups.html', {'mygroups': groups})
            else:
                groups = Group.objects.filter(date__icontains=search_id, talker_id=request.session['talker_id']).order_by('-id')
                if groups.count() > 0:
                    return render(request, 'may/mygroups.html', {'mygroups': groups})
                else:
                    return redirect('/mygroups')
    else:
        return redirect('/mygroups')

@csrf_protect
@csrf_exempt
def schallgroups(request):
    if request.method == 'POST':
        search_id = request.POST.get('q', None)
        groups = Group.objects.filter(talker_name__icontains=search_id).order_by('-id')
        if groups.count()>0:
            return render(request, 'may/groups.html', {'groups': groups})
        else:
            groups = Group.objects.filter(name__icontains=search_id).order_by('-id')
            if groups.count() > 0:
                return render(request, 'may/groups.html', {'groups': groups})
            else:
                groups = Group.objects.filter(date__icontains=search_id).order_by('-id')
                if groups.count() > 0:
                    return render(request, 'may/groups.html', {'groups': groups})
                else:
                    return redirect('/groups')
    else:
        return redirect('/groups')

@csrf_protect
@csrf_exempt
def schmynotes(request, user_id):
    if request.method == 'POST':
        search_id = request.POST.get('q', None)
        notes = Note.objects.filter(name__icontains=search_id, talker=request.session['talker_id']).order_by('-id')
        if notes.count()>0:
            noteList = []
            attendeeList = []
            for note in notes:
                if note.group != '':
                    attendees = Attendee.objects.filter(talker = user_id, note_id=note.pk)
                    noteList.insert(0, note)
                    attendeeList.insert(0, attendees)

            return render(request, 'may/transcripts.html',
                                  {'notes': noteList, 'attendees': attendeeList, 'user_id': user_id})

        else:
            notes = Note.objects.filter(email__icontains=search_id, talker=request.session['talker_id']).order_by('-id')
            if notes.count() > 0:
                noteList = []
                attendeeList = []
                for note in notes:
                    if note.group != '':
                        attendees = Attendee.objects.filter(talker = user_id, note_id=note.pk)
                        noteList.insert(0, note)
                        attendeeList.insert(0, attendees)

                return render(request, 'may/transcripts.html',
                                      {'notes': noteList, 'attendees': attendeeList, 'user_id': user_id})
            else:
                notes = Note.objects.filter(note__icontains=search_id, talker=request.session['talker_id']).order_by('-id')
                if notes.count() > 0:
                    noteList = []
                    attendeeList = []
                    for note in notes:
                        if note.group != '':
                            attendees = Attendee.objects.filter(talker = user_id, note_id=note.pk)
                            noteList.insert(0, note)
                            attendeeList.insert(0, attendees)

                    return render(request, 'may/transcripts.html',
                                          {'notes': noteList, 'attendees': attendeeList, 'user_id': user_id})
                else:
                    notes = Note.objects.filter(urgency__icontains=search_id, talker=request.session['talker_id']).order_by('-id')
                    if notes.count() > 0:
                        noteList = []
                        attendeeList = []
                        for note in notes:
                            if note.group != '':
                                attendees = Attendee.objects.filter(talker = user_id, note_id=note.pk)
                                noteList.insert(0, note)
                                attendeeList.insert(0, attendees)

                        return render(request, 'may/transcripts.html',
                                              {'notes': noteList, 'attendees': attendeeList, 'user_id': user_id})
                    else:
                        notes = Note.objects.filter(summary__icontains=search_id, talker=request.session['talker_id']).order_by('-id')
                        if notes.count() > 0:
                            noteList = []
                            attendeeList = []
                            for note in notes:
                                if note.group != '':
                                    attendees = Attendee.objects.filter(talker = user_id, note_id=note.pk)
                                    noteList.insert(0, note)
                                    attendeeList.insert(0, attendees)

                            return render(request, 'may/transcripts.html',
                                                  {'notes': noteList, 'attendees': attendeeList, 'user_id': user_id})
                        else:
                            notes = Note.objects.filter(callofaction__icontains=search_id, talker=request.session['talker_id']).order_by('-id')
                            if notes.count() > 0:
                                noteList = []
                                attendeeList = []
                                for note in notes:
                                    if note.group != '':
                                        attendees = Attendee.objects.filter(talker = user_id, note_id=note.pk)
                                        noteList.insert(0, note)
                                        attendeeList.insert(0, attendees)

                                return render(request, 'may/transcripts.html',
                                                      {'notes': noteList, 'attendees': attendeeList,
                                                       'user_id': user_id})
                            else:
                                notes = Note.objects.filter(email__icontains=search_id, talker=request.session['talker_id']).order_by('-id')
                                if notes.count() > 0:
                                    noteList = []
                                    attendeeList = []
                                    for note in notes:
                                        if note.group != '':
                                            attendees = Attendee.objects.filter(talker = user_id, note_id=note.pk)
                                            noteList.insert(0, note)
                                            attendeeList.insert(0, attendees)

                                    return render(request, 'may/transcripts.html',
                                                          {'notes': noteList, 'attendees': attendeeList,
                                                           'user_id': user_id})
                                else:
                                    allnotes = Note.objects.filter(talker=request.session['talker_id']).order_by('-id')
                                    noteList = []
                                    attendeeList = []
                                    for note in allnotes:
                                        # return HttpResponse(note.start_time)
                                        st = time.strftime('%Y-%m-%d %H:%M', time.gmtime(int(note.start_time)/1000.0))
                                        et = time.strftime('%Y-%m-%d %H:%M', time.gmtime(int(note.end_time) / 1000.0))
                                        if search_id in st:
                                            attendees = Attendee.objects.filter(talker = user_id, note_id=note.pk)
                                            noteList.insert(0, note)
                                            attendeeList.insert(0, attendees)
                                        elif search_id in et:
                                            attendees = Attendee.objects.filter(talker = user_id, note_id=note.pk)
                                            noteList.insert(0, note)
                                            attendeeList.insert(0, attendees)
                                    return render(request, 'may/transcripts.html',
                                                          {'notes': noteList, 'attendees': attendeeList,
                                                           'user_id': user_id})

    else:
        return redirect('/transcripts/' + user_id)

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def invitetogroup(request):
    if request.method == 'POST':
        group_id = request.POST.get('group_id', '')
        userids = request.POST.getlist('items[]')
        if len(userids) == 0:
            return render(request, 'may/result.html',
                          {'response': 'Please select at least a friend.'})
        group = Group.objects.get(id=group_id)
        message = ''
        try:
            message = request.POST.get('message', '')
        except MultiValueDictKeyError:
            print('No message')
        message = 'I invited you to ' + group.name + '(Created at ' + group.date + ').<br>' + message + '<br>Regards.<br><br>https://scwribe.com'

        html = """\
                        <html>
                          <head></head>
                          <body>
                          <a href="https://scwribe.com"><img src="https://scwribe.com/static/may/images/logo.png" style="width:90px;height:90px;border-radius: 8%; margin-left:25px;"/></a>
                            <h2 style="margin-left:10px; color:#02839a;">Scwribe friend's invitation</h2>
                            <div style="font-size:16px; word-break: break-all; word-wrap: break-word;">
                                {mes}
                            </div>
                          </body>
                        </html>
                        """
        html = html.format(mes=message)
        me = Talker.objects.get(id=request.session['talker_id'])
        fromaddress = me.email
        mailList = []
        useridList = []
        for i in userids:
            user = Talker.objects.get(id=i)
            mailList.append(user.email)
            useridList.append(i)

        msg = EmailMultiAlternatives('You have been invited by scwribe friend', '', fromaddress, mailList)
        msg.attach_alternative(html, "text/html")
        msg.send(fail_silently=False)

        talkers = []
        users = Talker.objects.all()
        me = Talker.objects.get(id=request.session['talker_id'])
        for user in users:
            if user.email != me.email:
                talkers.insert(0, user)
        groups = Group.objects.filter(talker_id=request.session['talker_id']).order_by('-id')
        return render(request, 'may/mygroups.html', {'mygroups': groups, 'users': talkers, 'me':me, 'userids':useridList, 'groupname':group.name, 'note':'success'})
    else:
        return redirect('/mygroups')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def userinvite(request):
    if request.method == 'POST':
        groupid = request.POST.get('groupid')
        userids = request.POST.getlist('userids[]')
        userids = userids[0].split(',')
        message = ''
        try:
            message = request.POST.get('message', '')
        except MultiValueDictKeyError:
            print('No message')
        me = Talker.objects.get(id=request.session['talker_id'])
        mailList = []
        useridList = []

        for i in userids:
            user = Talker.objects.get(id=int(i))
            mailList.append(user.email)
            useridList.append(i)
        group = Group.objects.get(id=int(groupid))
        message = 'I invited you to ' + group.name + '(Created at ' + group.date + ').<br>' + message + '<br>Regards.<br><br>https://scwribe.com'

        html = """\
                                            <html>
                                              <head></head>
                                              <body>
                                              <a href="https://scwribe.com"><img src="https://scwribe.com/static/may/images/logo.png" style="width:90px;height:90px;border-radius: 8%; margin-left:25px;"/></a>
                                                <h2 style="margin-left:10px; color:#02839a;">Scwribe friend's invitation</h2>
                                                <div style="font-size:16px; word-break: break-all; word-wrap: break-word;">
                                                    {mes}
                                                </div>
                                              </body>
                                            </html>
                                            """
        html = html.format(mes=message)
        fromaddress = me.email

        msg = EmailMultiAlternatives('You have been invited by scwribe friend', '', fromaddress, mailList)
        msg.attach_alternative(html, "text/html")
        msg.send(fail_silently=False)

        talkers = []
        users = Talker.objects.all()
        me = Talker.objects.get(id=request.session['talker_id'])
        for user in users:
            if user.email != me.email:
                talkers.insert(0, user)
        groups = Group.objects.filter(talker_id=request.session['talker_id']).order_by('-id')
        return render(request, 'may/users.html', {'users': talkers, 'groups': groups, 'me':me, 'userids':useridList, 'groupname':group.name, 'note':'success'})
    else:
        return redirect('/users')

def get_notifications(request):
    me = Talker.objects.get(id=request.session['talker_id'])
    return render(request, 'may/notification.html', {'me': me})

def chat_page(request, user_id):
    roomlink = ''
    me = Talker.objects.get(id=request.session['talker_id'])
    user = Talker.objects.get(id=user_id)
    rooms = TalkingRoom.objects.filter(talker1=user_id, talker2=request.session['talker_id'])
    if rooms.count() > 0:
        roomlink = rooms[0].roomlink
    else:
        room = TalkingRoom()
        room.talker1 = request.session['talker_id']
        room.talker2 = user_id
        room.roomlink = me.link
        room.save()
        roomlink = me.link
    context = {
        'sender': me,
        'user': user,
        'roomlink': roomlink
    }
    request.session['the'] = user_id
    return render(request, 'may/chattest.html', context)

def calllogin(request):
    me = Talker.objects.get(id=request.session['talker_id'])
    me.callreg = '1'
    me.save()
    return redirect('/home')

def createTalkingLink(request):
    me = Talker.objects.get(id=request.session['talker_id'])
    if me.callreg == '0':
        return redirect('/groupvideocall')
    else:
        return redirect('/home')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def createTalkingRoom(request):
    if request.method == 'POST':
        link = request.POST.get('link', '')
        me = Talker.objects.get(id=request.session['talker_id'])
        me.callreg = '1'
        me.link = link
        me.save()
        return redirect('/home')
    else:
        return redirect('/home')

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def forgotpassword(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        usrs = Talker.objects.filter(email=email)
        if usrs.count() == 0:
            return render(request, 'may/result.html',
                          {'response': 'You haven\'t been registered.'})

        fromaddress = 'info@caycompany.com'
        mailList = []
        mailList.append(email)

        message = 'You are allowed to reset your password from your request.<br>For it, please click this link to reset your password.<br><br>https://scwribe.com/resetpassword?email=' + email

        html = """\
                                            <html>
                                              <head></head>
                                              <body>
                                              <a href="https://scwribe.com"><img src="https://scwribe.com/static/may/images/logo.png" style="width:90px;height:90px;border-radius: 8%; margin-left:25px;"/></a>
                                                <h2 style="margin-left:10px; color:#02839a;">Scwribe User's Security Update Information</h2>
                                                <div style="font-size:16px; word-break: break-all; word-wrap: break-word;">
                                                    {mes}
                                                </div>
                                              </body>
                                            </html>
                                            """
        html = html.format(mes=message)

        msg = EmailMultiAlternatives('We allowed you to reset your password', '', fromaddress, mailList)
        msg.attach_alternative(html, "text/html")
        msg.send(fail_silently=False)

        return redirect('/home')
    else:
        return redirect('/home')

def resetpassword(request):
    email = request.GET['email']
    return render(request, 'may/resetpwd.html', {'email':email})

@csrf_protect
@csrf_exempt
@permission_classes((AllowAny,))
@api_view(['GET', 'POST'])
def rstpwd(request):
    if request.method == 'POST':
        email = request.POST.get('email', '')
        pwd = request.POST.get('password', '')
        repwd = request.POST.get('repassword', '')
        if pwd != repwd:
            return render(request, 'may/result.html',
                          {'response': 'The passwords you entered don\'t match each other. Enter the same passwords.'})

        usrs = Talker.objects.filter(email=email)
        if usrs.count() > 0:
            users = Member.objects.filter(email=email)
            if users.count() > 0:
                user = users[0]
                originalpwd = user.password
                user.password = pwd

                user1 = authenticate(username=email, password=originalpwd)
                if user1 is not None:
                    user1.password = pwd
                    user1.set_password(pwd)
                    user1.save()

                    user.save()

                    usr = usrs[0]
                    usr.password = pwd
                    usr.save()
            else:
                comusers = ComMember.objects.filter(email=email)
                if comusers.count() > 0:
                    comuser = comusers[0]
                    comuser.password = pwd
                    comuser.save()
            return redirect('/logout')
        else:
            return render(request, 'may/result.html',
                          {'response': 'You haven\'t been registered. Please register...'})
    else:
        return redirect('/logout')

def call(request):
    me = Talker.objects.get(id=request.session['talker_id'])
    return render(request, 'may/calltest2.html', {'me':me})

def video_call(request):
    # return render(request, 'may/call3.html')
    return render(request, 'may/sinchphonecall.html')

def groupvideocall(request):
    # return render(request, 'may/call3.html')
    return render(request, 'may/sinchgroupvideocall.html')











